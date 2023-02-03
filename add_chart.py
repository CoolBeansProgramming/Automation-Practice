from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# turn this file into an executable 
import os
import sys


application_path = os.path.dirname(sys.executable)

month = input('Enter the month: ')

# create input path (join application path with file)
input_path = os.path.join(application_path,'pivot_table.xlsx')


# load workbook/sheet with data
wb=load_workbook(input_path)
sheet=wb['Report']

#create and format titles
sheet['A1']= 'Sales Report' # title
sheet['A1'].font=Font('Arial', bold=True, size=20)

sheet['A2']=f'{month} 2023' # subtitle
sheet['A2'].font=Font('Times New Roman', bold=True, size=12)


# get active min/max columns
min_column=wb.active.min_column
max_column=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row

# get the letters of the columns
for i in range(min_column+1,max_column+1):
    letter=get_column_letter(i)
    sheet[f'{letter}{max_row+1}']=f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style='Currency'

# create a barchar  object; format it 
barchart = BarChart()

# find the data and categories
data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row) # go across the rows
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) # always on the left of the sheet 

# add the data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# make the chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 2


# create output path
output_path = os.path.join(application_path, f'report_{month}.xlsx')

# save bar graph
wb.save(output_path)

